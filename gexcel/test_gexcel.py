"""
TIME: 2019/11/24 12:01
Author: Guan
GitHub: https://github.com/youguanxinqing
EMAIL: youguanxinqing@qq.com
"""
import os
import hashlib
import unittest

import openpyxl

from .fromexcel import FromExcel
from .toexcel import ToExcel


class TestFromExcel(unittest.TestCase):
    def setUp(self):
        m = hashlib.md5("demo".encode("utf-8"))
        self.filename = f"{m.hexdigest()}.xlsx"
        wb = openpyxl.Workbook()
        sheet = wb.active
        # 制作模拟数据
        [[sheet.cell(row, col, f"({row}, {col})") for col in range(1, 21)] for row in range(1, 21)]
        wb.save(self.filename)
        wb.close()

        self.fe = FromExcel(f=self.filename)

    def tearDown(self):
        self.fe.close()
        if os.path.exists(self.filename):
            os.remove(self.filename)

    def test_read_point(self):
        """
        测试单元格读取
        """
        res = self.fe.read_point(3, "A")
        self.assertEqual(res, "(3, 1)")

        res = self.fe.read_point(100, "Z")
        self.assertEqual(res, None)

    def test_read_line(self):
        row, num = 10, 5
        res = self.fe.read_line(row, num)
        expected_res = [f"{(10, i)}" for i in range(1, num+1)]
        self.assertEqual(expected_res, res)

        col_start = 5
        res = self.fe.read_line(row, num, 5)
        expected_res = [f"{(10, i)}" for i in range(col_start, num+col_start)]
        self.assertEqual(expected_res, res)

    def test_read_rect(self):
        lt, lb, rt, rb = 3, 10, "B", "J"
        res = self.fe.read_rect(lt, lb, rt, rb)

        j_start = FromExcel._alpha_to_digit(rt)
        j_end = FromExcel._alpha_to_digit(rb)
        expected_res = [
            [f"{(i, j)}" for j in range(j_start, j_end+1)] for i in range(lt, lb+1)
        ]
        self.assertEqual(expected_res, res)

    def test_read_page(self):
        res = self.fe.read_page(num=20)
        expect_res = [[f"({row}, {col})" for col in range(1, 21)] for row in range(1, 21)]
        self.assertEqual(expect_res, list(res))


class TestToExcel(unittest.TestCase):
    def setUp(self):
        m = hashlib.md5("demo".encode("utf-8"))
        self.filename = f"{m.hexdigest()}.xlsx"

        self.te = ToExcel(".", self.filename)

    def tearDown(self):
        self.te.close()
        if os.path.exists(self.filename):
            os.remove(self.filename)

    def test_locate_point(self):
        self.te.locate_point = (1, "A", "ZTY")
        self.te.locate_point = (10, "J", "Ying")
        self.te.save().close()

        wb = openpyxl.load_workbook(self.filename, read_only=True)
        sheet = wb.active
        res_a = sheet.cell(1, 1).value
        res_b = sheet.cell(10, 10).value
        res_c = sheet.cell(8, 19).value
        wb.close()

        self.assertEqual("ZTY", res_a)
        self.assertEqual("Ying", res_b)
        self.assertNotEqual("Ying", res_c)

    def test_locate_line(self):
        cur_row = self.te._row
        self.te.locate_line = (("z", "t", "y"),)
        self.te.save().close()

        wb = openpyxl.load_workbook(self.filename, read_only=True)
        sheet = wb.active
        res_a = sheet.cell(cur_row, 1).value
        res_b = sheet.cell(cur_row, 2).value
        res_c = sheet.cell(cur_row, 3).value
        wb.close()

        self.assertEqual(("z", "t", "y"), (res_a, res_b, res_c))

    def test_locate_line_row(self):
        """
        指定行号写入行内容
        """
        row = 10
        self.te.locate_line = (row, ("z", "t", "y"))
        self.te.save().close()

        wb = openpyxl.load_workbook(self.filename, read_only=True)
        sheet = wb.active
        res_a = sheet.cell(row, 1).value
        res_b = sheet.cell(row, 2).value
        res_c = sheet.cell(row, 3).value
        wb.close()

        self.assertEqual(("z", "t", "y"), (res_a, res_b, res_c))

    def test_locate_line_row_col(self):
        row, col = 15, 10
        self.te.locate_line = (row, col, ("z", "t", "y"))
        self.te.save().close()

        wb = openpyxl.load_workbook(self.filename, read_only=True)
        sheet = wb.active
        res_a = sheet.cell(row, col).value
        res_b = sheet.cell(row, col+1).value
        res_c = sheet.cell(row, col+2).value
        wb.close()

        self.assertEqual(("z", "t", "y"), (res_a, res_b, res_c))
